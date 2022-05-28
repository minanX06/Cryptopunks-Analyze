#과학 수행평가 때 사용한 파이썬 프로그래밍 코드입니다.
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

crypto = []
cryp2 = []
req = requests.get("https://cryptopunks.app/cryptopunks/attributes")
soup = BeautifulSoup(req.text, "html.parser")
for anchor in soup.select("tr.center.attribute-row"):
    crypto.append(str(anchor))

for i in range(5):
    crypto.pop(0)

for k in range(0, 87):
    att = crypto[k].split("<td>")[2:5]

    for j in range(0, 3):
        cryp2.append(att[j].split("\n")[1].strip())

wb = Workbook()
ws = wb.active

for i in range(87):
    ws["A{}".format(i + 1)] = cryp2[i*3]
    ws["B{}".format(i + 1)] = cryp2[i*3 + 2][:-1]

wb.save("cryptopunks-analyze.xlsx")
wb.close()
